from flask import Flask, request, send_file, jsonify
from notion_client import Client as NotionClient
import requests
import os
import io
import tempfile
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment
import smtplib
from email.message import EmailMessage
from PIL import Image
import time
from collections import OrderedDict
from datetime import datetime, timezone
from dateutil import parser
import uuid
import logging
from apscheduler.schedulers.background import BackgroundScheduler

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("brandid_processor.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger("BrandIDProcessor")

app = Flask(__name__)

# === ENVIRONMENT VARIABLES ===
NOTION_TOKEN = os.getenv("NOTION_TOKEN")
DATABASE_ID = os.getenv("NOTION_DATABASE_ID")
TEMPLATE_PATH = "invoice-watermarked.xlsx"
SCHEDULER_INTERVAL = int(os.getenv("SCHEDULER_INTERVAL", "60"))  # Default to 60 minutes

# Initialize Notion client
notion = NotionClient(auth=NOTION_TOKEN)

# Simple in-memory cache of processed events
# Using OrderedDict to limit memory usage (keeps only most recent 100 events)
PROCESSED_EVENTS = OrderedDict()
MAX_CACHE_SIZE = 100

def get_property_value(properties, name, type_name):
    """Extract values from Notion property objects"""
    if name not in properties:
        return ""
        
    prop = properties[name]
    
    if type_name == "title" and "title" in prop:
        return prop["title"][0]["plain_text"] if prop["title"] else ""
    elif type_name == "rich_text" and "rich_text" in prop:
        return prop["rich_text"][0]["plain_text"] if prop["rich_text"] else ""
    elif type_name == "email":
        return prop.get("email", "")
    elif type_name == "phone_number":
        return prop.get("phone_number", "")
    elif type_name == "checkbox":
        return prop.get("checkbox", False)
    
    return ""
      
# === BRAND ID GENERATION ===
def generate_brand_id(business_name, email):
    """Generate a unique Brand ID following these specific rules:
    
    Format: BRAND-{4-char name code}-{4-digit email code}
    
    Name code algorithm:
    1. For each word, take the first letter to build an initial string
    2. For each word, find first consonant and add after its initial
    3. Limit to 4 characters
    4. Don't repeat letters (use Y instead of X if a second X is needed)
    5. If no consonants, just take the left 4 chars of company name
    
    Examples:
    - "Acme Design Studio" → ACDS (A + C + D + S)
    - "Quick Shop" → QCSH (Q + C + S + H)
    - "AB" → ABXY (A + X + B + Y)
    - "Professional Services" → PRSV (P + R + S + V)
    - "AEIOU" → AEIO (just first 4 chars because no consonants)
    - "XYZ Corp" → XYCR (X + Y + C + R)
    """
    if not business_name or not isinstance(business_name, str) or business_name.strip() == "":
        business_name = "Unknown"
    
    if not email or not isinstance(email, str):
        email = "example@example.com"
    
    # Define vowels for consonant detection
    vowels = ['a', 'e', 'i', 'o', 'u']
    
    # Split into words and filter out empty strings
    words = [word for word in business_name.split() if word]
    if not words:
        words = ["Unknown"]
    
    # Check if there are any consonants in the entire name
    has_consonants = False
    for char in business_name:
        if char.isalpha() and char.lower() not in vowels:
            has_consonants = True
            break
    
    # If no consonants, just take the first 4 characters of the business name
    if not has_consonants:
        name_part = business_name[:4].upper()
        # Pad with X, Y, Z if needed
        if len(name_part) < 4:
            padding_chars = ['X', 'Y', 'Z']
            i = 0
            while len(name_part) < 4:
                name_part += padding_chars[i % 3]
                i += 1
    else:
        # Initialize result and used letters set
        name_part = ""
        used_letters = set()
        
        # Alternate adding initial and first consonant for each word
        for word in words:
            if len(name_part) >= 4:
                break
            
            # Add the initial (first letter)
            initial = word[0].upper()
            if initial not in used_letters:
                name_part += initial
                used_letters.add(initial)
            
            if len(name_part) >= 4:
                break
                
            # Find first consonant in the word
            consonant_found = False
            for char in word:
                if char.isalpha() and char.lower() not in vowels and char.upper() not in used_letters:
                    name_part += char.upper()
                    used_letters.add(char.upper())
                    consonant_found = True
                    break
            
            # If no unused consonant found, add an unused padding character
            if not consonant_found and len(name_part) < 4:
                for padding_char in ['X', 'Y', 'Z']:
                    if padding_char not in used_letters:
                        name_part += padding_char
                        used_letters.add(padding_char)
                        break
        
        # Trim to 4 characters if longer
        name_part = name_part[:4]
        
        # Pad with unused characters if shorter than 4
        padding_chars = ['X', 'Y', 'Z']
        i = 0
        while len(name_part) < 4:
            padding_char = padding_chars[i % 3]
            if padding_char not in used_letters:
                name_part += padding_char
                used_letters.add(padding_char)
            i += 1
    
    # Calculate ASCII value of email address
    email_ascii_sum = sum(ord(c) for c in email)
    email_part = str(email_ascii_sum)[-4:].zfill(4)  # Get rightmost 4 digits
    
    # Construct final Brand ID
    brand_id = f"BRAND-{name_part}-{email_part}"
    
    logger.info(f"Generated Brand ID for {business_name}: {brand_id}")
    return brand_id

# === UTILITIES ===
def send_email(recipient_email, subject, body, attachment_paths, business_name='', brand_id=''):
    """Send email with attachments and formatted HTML body"""
    smtp_server = os.getenv('SMTP_SERVER', 'smtp.gmail.com')
    smtp_port = int(os.getenv('SMTP_PORT', 587))
    smtp_user = os.getenv('SMTP_USER')
    smtp_pass = os.getenv('SMTP_PASS')
    sender_name = os.getenv('SENDER_NAME', 'Invoice Generator')

    # Create a more sophisticated email
    msg = EmailMessage()
    
    # Better subject line with personalization
    personalized_subject = f"{business_name} - {subject}" if business_name else subject
    msg['Subject'] = personalized_subject
    
    # Add proper From header with sender name
    msg['From'] = f'"{sender_name}" <{smtp_user}>'
    msg['To'] = recipient_email
    
    # Add more headers to improve deliverability
    # Add a unique Message-ID
    domain = smtp_user.split('@')[-1]
    msg['Message-ID'] = f"<{uuid.uuid4()}@{domain}>"
    
    # Add Date header
    msg['Date'] = datetime.now().strftime("%a, %d %b %Y %H:%M:%S %z")
    
    # Add X-Mailer header 
    msg['X-Mailer'] = 'InvoiceCustomizer Service'
    
    # Add a List-Unsubscribe header (helps with spam prevention)
    msg['List-Unsubscribe'] = f'<mailto:{smtp_user}?subject=Unsubscribe>'
    
    # Create personalized HTML content
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1">
        <title>{personalized_subject}</title>
        <style>
            body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; padding: 20px; }}
            .container {{ max-width: 600px; margin: 0 auto; }}
            .header {{ background-color: #f8f9fa; padding: 15px; border-bottom: 1px solid #e9ecef; }}
            .content {{ padding: 20px 0; }}
            .footer {{ font-size: 12px; color: #6c757d; padding-top: 20px; border-top: 1px solid #e9ecef; }}
            .brand-id-box {{ background-color: #f0f8ff; border: 1px solid #b0c4de; padding: 15px; border-radius: 5px; margin: 20px 0; text-align: center; }}
            .brand-id {{ font-size: 24px; font-weight: bold; letter-spacing: 1px; color: #0056b3; }}
            .instructions {{ background-color: #fffaf0; border-left: 4px solid #ffa500; padding: 10px; margin: 15px 0; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h2>Your Custom Invoice Template & Brand ID</h2>
            </div>
            <div class="content">
                <p>Hello{' ' + business_name if business_name else ''},</p>
                
                <p>Thank you for using our Invoice Generator service! Your custom Excel invoice template is now ready.</p>
                
                <div class="brand-id-box">
                    <p>Your unique Brand ID is:</p>
                    <p class="brand-id">{brand_id}</p>
                </div>
                
                <div class="instructions">
                    <p><strong>Important:</strong> Save this Brand ID for future purchases. When you buy additional templates on Etsy, include this Brand ID in your order notes to have them automatically customized with your business information.</p>
                </div>
                
                <p>We've attached the spreadsheet to this email. You can fill in the item details, and the calculations will be performed automatically.</p>
                
                <p>This customized invoice includes:</p>
                <ul>
                    <li>Your business information and logo</li>
                    <li>Customized tax rate</li>
                    <li>Protected formulas to prevent accidental changes</li>
                </ul>
                
                <p>If you have any questions or need assistance, please reply to this email.</p>
            </div>
            <div class="footer">
                <p>This is a transactional email sent to you because you submitted the Invoice Customization Form.</p>
                <p>To unsubscribe from future emails, please reply with "Unsubscribe" in the subject line.</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    # Set both plain text and HTML content
    msg.set_content(body)
    msg.add_alternative(html_content, subtype='html')

    # Add attachments
    for attachment_path in attachment_paths:
        with open(attachment_path, 'rb') as f:
            file_data = f.read()
            file_name = os.path.basename(attachment_path)

        if attachment_path.endswith('.xlsx'):
            maintype = 'application'
            subtype = 'vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        else:
            maintype = 'application'
            subtype = 'octet-stream'

        msg.add_attachment(file_data, maintype=maintype, subtype=subtype, filename=file_name)

    # Send the email with a retry mechanism
    max_retries = 3
    retry_delay = 2  # seconds
    
    for attempt in range(max_retries):
        try:
            with smtplib.SMTP(smtp_server, smtp_port) as server:
                server.starttls()
                server.login(smtp_user, smtp_pass)
                server.send_message(msg)
                logger.info(f"✅ Email sent successfully to {recipient_email}")
                return True
        except Exception as e:
            if attempt < max_retries - 1:
                logger.warning(f"⚠️ Email attempt {attempt + 1} failed. Retrying in {retry_delay} seconds: {e}")
                time.sleep(retry_delay)
                retry_delay *= 2  # Exponential backoff
            else:
                logger.error(f"❌ All email attempts failed: {e}")
                return False

def remove_background(image_file, tolerance=50):
    """Remove background from logo image"""
    img = Image.open(image_file).convert("RGBA")
    datas = img.getdata()

    new_data = []
    background_color = datas[0]
    logger.debug(f"Detected background color: {background_color}")
    
    for item in datas:
        if all(abs(item[i] - background_color[i]) <= tolerance for i in range(3)):
            new_data.append((255, 255, 255, 0))
        else:
            new_data.append(item)
    img.putdata(new_data)
    return img

def update_notion_database(fields, event_id=None):
    """Updates the Notion database with preview information using existing fields"""
    if not NOTION_TOKEN or not DATABASE_ID:
        logger.warning("⚠️ Notion credentials not set, skipping database update")
        return
    
    try:
        # Extract relevant information
        business_name = fields.get('Company Name', 'Unknown')
        email = fields.get('Email', '')
        timestamp = datetime.now().isoformat()
        
        # Generate a unique Brand ID
        brand_id = ""
        if business_name and email:
            brand_id = generate_brand_id(business_name, email)
            fields['BrandID'] = brand_id
        
        # Prepare the data for Notion using only existing fields
        properties = {
            "Name": {"title": [{"text": {"content": business_name}}]},
            "Email": {"email": email},
            "Timestamp": {"date": {"start": timestamp}},
            "Validated": {"checkbox": True},
            "Email Sent": {"checkbox": False}  # Will be updated to True when email is sent
        }
        
        # Add Brand ID to properties
        if brand_id:
            properties["BrandID"] = {"rich_text": [{"text": {"content": brand_id}}]}
        
        # If we have a company name, add it to the Company field
        if fields.get('Company Name'):
            properties["Company"] = {"rich_text": [{"text": {"content": fields.get('Company Name', '')}}]}
        
        # Check if this email already exists in the database
        existing_pages = notion.databases.query(
            database_id=DATABASE_ID,
            filter={
                "property": "Email",
                "email": {
                    "equals": email
                }
            }
        ).get("results", [])
        
        if existing_pages:
            # Update existing page
            notion.pages.update(
                page_id=existing_pages[0]["id"],
                properties=properties
            )
            page_id = existing_pages[0]["id"]
            logger.info(f"✅ Updated existing Notion entry for {business_name}")
        else:
            # Create new page
            response = notion.pages.create(
                parent={"database_id": DATABASE_ID},
                properties=properties
            )
            page_id = response["id"]
            logger.info(f"✅ Created new Notion entry for {business_name}")
        
        return page_id
            
    except Exception as e:
        logger.error(f"⚠️ Error updating Notion database: {e}")
        return None
        
def protect_workbook(workbook, password='etsysc123'):
    """Apply protection to Excel workbook to prevent accidental changes"""
    for sheet in workbook.worksheets:
        # Enable protection with specific options
        sheet.protection.sheet = True
        sheet.protection.password = password
        
        # Disable object editing/deletion
        sheet.protection.objects = True
        sheet.protection.scenarios = True
        
        # Other protection options
        sheet.protection.selectLockedCells = False
        sheet.protection.selectUnlockedCells = False
        sheet.protection.formatCells = False
        sheet.protection.formatColumns = False
        sheet.protection.formatRows = False
        sheet.protection.insertColumns = False
        sheet.protection.insertRows = False
        sheet.protection.insertHyperlinks = False
        sheet.protection.deleteColumns = False
        sheet.protection.deleteRows = False
        sheet.protection.sort = False
        sheet.protection.autoFilter = False
        sheet.protection.pivotTables = False
        sheet.protection.drawings = True  # Specifically protects drawings/images

def insert_logo(ws, image_bytes):
    """Insert logo into worksheet"""
    # Create a temporary file that won't be deleted until program exit
    temp_logo = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
    temp_logo.write(image_bytes)
    temp_logo.close()  # Close the file but don't delete it

    try:
        img = OpenpyxlImage(temp_logo.name)
        col_width = ws.column_dimensions['A'].width or 8
        row_height = ws.row_dimensions[1].height or 15

        img.width = int(col_width * 7.5)
        img.height = int(row_height * 1.33)
        img.anchor = 'A1'

        ws.add_image(img)
        
        # Return the filename so we can delete it after the workbook is saved
        return temp_logo.name
    except Exception as e:
        logger.error(f"⚠️ Error adding logo: {e}")
        # If there's an error, try to clean up the file
        try:
            os.unlink(temp_logo.name)
        except:
            pass
        return None

def insert_watermark_background(ws):
    """Insert watermark into worksheet"""
    if os.path.exists("watermark.png"):
        with open("watermark.png", 'rb') as img_file:
            ws._background = img_file.read()

def is_stale_event(event_time, max_age_minutes=5):
    """Check if an event is too old to process"""
    try:
        event_datetime = parser.parse(event_time)
        current_time = datetime.now(timezone.utc)
        time_diff = (current_time - event_datetime).total_seconds() / 60
        return time_diff > max_age_minutes, time_diff
    except Exception as e:
        logger.error(f"⚠️ Error parsing event time: {e}")
        return False, 0  # If we can't parse the time, assume it's not stale

def extract_notion_properties(page):
    """Extract relevant properties from a Notion page"""
    properties = page.get("properties", {})
    fields = {}
    
    # Extract Company Name
    title_property = next((prop_name for prop_name, prop in properties.items() 
                         if prop.get("type") == "title"), None)
    if title_property:
        title_objects = properties[title_property].get("title", [])
        if title_objects:
            fields["Company Name"] = title_objects[0].get("plain_text", "")
    
    # Extract Email
    email_property = next((prop_name for prop_name, prop in properties.items() 
                         if prop.get("type") == "email"), None)
    if email_property:
        fields["Email"] = properties[email_property].get("email", "")
    
    # Extract Logo URL
    for prop_name, prop in properties.items():
        if prop.get("type") == "url" and "logo" in prop_name.lower():
            fields["Logo URL"] = prop.get("url", "")
        elif prop.get("type") == "rich_text" and "logo" in prop_name.lower():
            rich_texts = prop.get("rich_text", [])
            if rich_texts:
                fields["Logo URL"] = rich_texts[0].get("plain_text", "")
    
    # Extract Address
    for prop_name, prop in properties.items():
        if prop.get("type") == "rich_text" and "address" in prop_name.lower():
            rich_texts = prop.get("rich_text", [])
            if rich_texts:
                fields["Address"] = rich_texts[0].get("plain_text", "")
    
    # Extract City, State ZIP
    for prop_name, prop in properties.items():
        if prop.get("type") == "rich_text" and ("city" in prop_name.lower() or "zip" in prop_name.lower()):
            rich_texts = prop.get("rich_text", [])
            if rich_texts:
                fields["City, State ZIP"] = rich_texts[0].get("plain_text", "")
    
    # Extract Phone
    for prop_name, prop in properties.items():
        if prop.get("type") == "phone_number" or (prop.get("type") == "rich_text" and "phone" in prop_name.lower()):
            if prop.get("type") == "phone_number":
                fields["Phone"] = prop.get("phone_number", "")
            else:
                rich_texts = prop.get("rich_text", [])
                if rich_texts:
                    fields["Phone"] = rich_texts[0].get("plain_text", "")
    
    # Extract Tax %
    for prop_name, prop in properties.items():
        if prop.get("type") == "number" and "tax" in prop_name.lower():
            fields["Tax %"] = str(prop.get("number", 7))
    
    # Extract Currency
    for prop_name, prop in properties.items():
        if prop.get("type") == "select" and "currency" in prop_name.lower():
            select = prop.get("select", {})
            fields["Currency"] = select.get("name", "USD")
        elif prop.get("type") == "rich_text" and "currency" in prop_name.lower():
            rich_texts = prop.get("rich_text", [])
            if rich_texts:
                fields["Currency"] = rich_texts[0].get("plain_text", "USD")
    
    # Extract BrandID if it exists
    for prop_name, prop in properties.items():
        if prop.get("type") == "rich_text" and "brandid" in prop_name.lower():
            rich_texts = prop.get("rich_text", [])
            if rich_texts:
                fields["BrandID"] = rich_texts[0].get("plain_text", "")
    
    return fields

def process_template(fields, page_id):
    """Generate a customized template based on customer fields and Notion page ID"""
    temp_files = []  # List to keep track of temporary files to clean up
    
    try:
        business_name = fields.get("Company Name", "Your Business")
        address1 = fields.get("Address", "")
        address2 = fields.get("City, State ZIP", "")
        phone = fields.get("Phone", "")
        email = fields.get("Email", "")
        tax_percentage = fields.get("Tax %", "7")
        currency = fields.get("Currency", "USD")
        
        # Get logo directly from Notion page
        logo_bytes = get_logo_from_notion(page_id)
        
        if not logo_bytes:
            logger.warning(f"No logo found for {business_name}, proceeding without logo")
            # Continue with template generation without logo
        else:
            # Process the logo to remove background if logo exists
            try:
                processed_logo = remove_background(io.BytesIO(logo_bytes))
                logo_bytes_io = io.BytesIO()
                processed_logo.save(logo_bytes_io, format="PNG")
                logo_bytes_io.seek(0)
                logo_bytes = logo_bytes_io.read()
            except Exception as e:
                logger.error(f"Error processing logo: {e}")
                # Continue with the original logo if processing fails
                logo_bytes_io = io.BytesIO(logo_bytes)
        
        # Load template and customize
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        ws = wb.active
        
        insert_watermark_background(ws)
        
        # Insert business information
        ws['A2'] = business_name
        ws['A3'] = address1
        ws['A4'] = address2
        ws['A5'] = phone
        ws['A6'] = email
        ws['D30'] = f"Tax ({tax_percentage}%)"
        ws['E30'] = f'=IF(NOT(IsGoogleSheets),E29*{float(tax_percentage)}/100,"GOOGLE SHEETS DETECTED")'
        ws['C32'] = f"All amounts shown in {currency}"
        ws.merge_cells('C32:E32')
        ws['C32'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Insert logo if available
        if logo_bytes:
            logo_temp_path = insert_logo(ws, logo_bytes)
            if logo_temp_path:
                temp_files.append(logo_temp_path)
        
        # Apply protection
        protect_workbook(wb)
        
        # Save to temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            temp_files.append(tmp.name)
            return tmp.name, temp_files
    
    except Exception as e:
        logger.error(f"Error processing template: {e}")
        return None, temp_files

def get_logo_from_notion(page_id):
    """Extract logo image directly from Notion page blocks"""
    try:
        blocks = notion.blocks.children.list(block_id=page_id).get("results", [])
        
        for block in blocks:
            if block["type"] == "image":
                image_block = block["image"]
                if image_block["type"] == "file":
                    image_url = image_block["file"]["url"]
                    try:
                        response = requests.get(image_url)
                        if response.status_code == 200:
                            return response.content
                    except Exception as e:
                        logger.error(f"Error downloading logo from Notion URL: {e}")
                elif image_block["type"] == "external":
                    image_url = image_block["external"]["url"]
                    try:
                        response = requests.get(image_url)
                        if response.status_code == 200:
                            return response.content
                    except Exception as e:
                        logger.error(f"Error downloading logo from external URL: {e}")
        
        # Check for logo in properties
        page = notion.pages.retrieve(page_id=page_id)
        if "properties" in page:
            for prop_name, prop_data in page["properties"].items():
                if prop_data.get("type") == "files" and prop_data.get("files"):
                    for file in prop_data["files"]:
                        if file["type"] == "file":
                            image_url = file["file"]["url"]
                            try:
                                response = requests.get(image_url)
                                if response.status_code == 200:
                                    return response.content
                            except Exception as e:
                                logger.error(f"Error downloading logo from property: {e}")
        
        logger.warning(f"No logo found in page {page_id}")
        return None
    except Exception as e:
        logger.error(f"Error retrieving logo from Notion: {e}")
        return None

def update_notion_with_brand_id(page_id, brand_id, email_sent=False):
    """Update Notion record with Brand ID and email status"""
    try:
        properties = {
            "BrandID": {"rich_text": [{"text": {"content": brand_id}}]}
        }
        
        if email_sent:
            properties["Email Sent"] = {"checkbox": True}
        
        notion.pages.update(
            page_id=page_id,
            properties=properties
        )
        logger.info(f"✅ Updated Notion page {page_id} with Brand ID {brand_id}")
        return True
    except Exception as e:
        logger.error(f"❌ Failed to update Notion page {page_id}: {e}")
        return False

# === PROCESSING FUNCTION ===
def process_pending_records():
    """
    Main function to scan Notion database for records that need processing:
    - Records without a Brand ID
    - Records with a Brand ID but where email wasn't sent successfully
    """
    logger.info("Scanning for records that need processing...")
    
    if not NOTION_TOKEN or not DATABASE_ID:
        logger.error("Notion credentials not set, cannot process records")
        return
    
    try:
        # Query Notion for records that either:
        # 1. Don't have a Brand ID, OR
        # 2. Have a Brand ID but email wasn't sent (Email Sent is false)
        results = notion.databases.query(
            database_id=DATABASE_ID,
            filter={
                "or": [
                    {
                        "property": "BrandID", 
                        "rich_text": {
                            "is_empty": True
                        }
                    },
                    {
                        "and": [
                            {
                                "property": "BrandID",
                                "rich_text": {
                                    "is_not_empty": True
                                }
                            },
                            {
                                "property": "Email Sent",
                                "checkbox": {
                                    "equals": False
                                }
                            }
                        ]
                    }
                ]
            }
        ).get("results", [])
        
        logger.info(f"Found {len(results)} records that need processing")
        
        # Process each record
        processed_count = 0
        for record in results:
            try:
                page_id = record["id"]
                properties = record["properties"]
                
                # Extract business information
                business_name =  get_property_value(properties, "Company", "title")
                email = get_property_value(properties, "Email", "email")
                phone = get_property_value(properties, "Phone", "phone_number") 
                address = get_property_value(properties, "Address", "rich_text")
                city_state_zip = get_property_value(properties, "CityStateZip", "rich_text")
                tax_percentage = get_property_value(properties, "Tax Percentage", "number") or "7"
                
                if not business_name or not email:
                    logger.warning(f"Skipping record {page_id}: Missing required fields (business name or email)")
                    continue
                
                # Check if record already has a Brand ID
                existing_brand_id = get_property_value(properties, "BrandID", "rich_text")
                
                # Generate or use existing Brand ID
                if existing_brand_id:
                    brand_id = existing_brand_id
                    logger.info(f"Using existing Brand ID {brand_id} for {business_name}")
                else:
                    brand_id = generate_brand_id(business_name)
                    logger.info(f"Generated Brand ID {brand_id} for {business_name}")
                
                # Collect business data for invoice generation
                fields = {
                    "Company Name": business_name,
                    "Address": address,
                    "City, State ZIP": city_state_zip,
                    "Phone": phone,
                    "Email": email,
                    "Tax %": "7",  # Default
                    "Currency": "USD"  # Default
                }
                
                # Generate invoice preview
                invoice_path, temp_files = process_template(fields, page_id)
                
                # Clean up temp files at the end
                try:
                    if invoice_path:
                        # Send email with Brand ID and preview
                        email_sent = send_email(
                            recipient_email=email,
                            subject="Your Custom Invoice Template & Brand ID",
                            body=brand_id,
                            attachment_paths=[invoice_path],
                            business_name=business_name
                        )
                        
                        # Update Notion with Brand ID if email was sent
                        if email_sent:
                            notion.pages.update(
                                page_id=page_id,
                                properties={
                                    "BrandID": {
                                        "rich_text": [
                                            {
                                                "text": {
                                                    "content": brand_id
                                                }
                                            }
                                        ]
                                    },
                                    "Email Sent": {
                                        "checkbox": True
                                    }
                                }
                            )
                            processed_count += 1
                            logger.info(f"Successfully processed {business_name} with Brand ID: {brand_id}")
                    else:
                        logger.error(f"Failed to generate template for {business_name}")
                finally:
                    # Clean up temporary files
                    for file_path in temp_files:
                        if file_path and os.path.exists(file_path):
                            os.unlink(file_path)
                
                # Small delay to avoid rate limits
                time.sleep(1)
                
            except Exception as e:
                logger.error(f"Error processing record {page_id if 'page_id' in locals() else 'unknown'}: {e}")
                continue
        
        logger.info(f"Successfully processed {processed_count} out of {len(results)} records")
        return processed_count
        
    except Exception as e:
        logger.error(f"Error querying Notion database: {e}")
        return 0

# === SCHEDULER ===
def start_scheduler():
    """Start the background scheduler"""
    scheduler = BackgroundScheduler()
    scheduler.add_job(
        process_pending_records, 
        'interval', 
        minutes=SCHEDULER_INTERVAL,
        id='process_pending_records_job'
    )
    scheduler.start()
    logger.info(f"Scheduler started, will run every {SCHEDULER_INTERVAL} minutes")
    return scheduler

# === WEBHOOK ===
@app.route("/preview_webhook", methods=["POST"])
def handle_preview_request():
    temp_files = []  # List to keep track of temporary files to clean up
    event_id = None
    
    try:
        data = request.json
        logger.info("🚀 Raw incoming data from Tally:", data)

        # Extract event ID and creation time for idempotency check
        event_id = data.get('eventId')
        event_time = data.get('createdAt')

        # Check if we've already processed this event SUCCESSFULLY
        if event_id in PROCESSED_EVENTS and PROCESSED_EVENTS[event_id].get("processed", False):
            logger.info(f"✅ Skipping already processed event: {event_id}")
            return '', 204
        
        # If we received it before but didn't process it successfully, we'll try again
        if event_id in PROCESSED_EVENTS:
            logger.warning(f"⚠️ Retrying previously failed event: {event_id}")
        
        # Check if the event is too old (stale)
        is_stale, time_diff = is_stale_event(event_time)
        if is_stale:
            logger.warning(f"⚠️ Skipping stale event from {time_diff:.1f} minutes ago: {event_id}")
            
            # Mark as received but not successfully processed
            if len(PROCESSED_EVENTS) >= MAX_CACHE_SIZE:
                PROCESSED_EVENTS.popitem(last=False)
            PROCESSED_EVENTS[event_id] = {"timestamp": time.time(), "processed": False}
            
            return '', 204

        # Extract fields from the webhook payload
        fields_list = data.get('data', {}).get('fields', [])
        fields = {}

        for field in fields_list:
            label = field.get('label')
            value = field.get('value')

            if isinstance(value, list):
                if value and isinstance(value[0], dict) and 'url' in value[0]:
                    value = value[0]['url']
                else:
                    value = None

            fields[label] = value

        business_name = fields.get('Company Name', 'Your Business')
        address1 = fields.get('Address', '')
        address2 = fields.get('City, State ZIP', '')
        phone = fields.get('Phone', '')
        email = fields.get('Email', '')
        tax_percentage = fields.get('Tax %', '7')
        currency = fields.get('Currency', 'USD')
        
        #logo_bytes = get
        if logo_bytes:
            # Process the logo to remove background
            processed_logo_bytes = remove_background(logo_bytes)
            
            # Insert logo
            logo_temp_path = insert_logo(ws, processed_logo_bytes)
            if logo_temp_path:
                temp_files.append(logo_temp_path)
        else:
            logger.warning(f"Skipping logo for {business_name} as no logo was found")
        
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        ws = wb.active

        insert_watermark_background(ws)

        ws['A2'] = business_name
        ws['A3'] = address1
        ws['A4'] = address2
        ws['A5'] = phone
        ws['A6'] = email
        ws['D30'] = f"Tax ({tax_percentage}%)"
        ws['E30'] = f'=IF(NOT(IsGoogleSheets),E29*{float(tax_percentage)}/100,"GOOGLE SHEETS DETECTED")'
        ws['C32'] = f"All amounts shown in {currency}"
        ws.merge_cells('C32:E32')
        ws['C32'].alignment = Alignment(horizontal='center', vertical='center')

        # Insert logo and get the temp file path
        logo_temp_path = insert_logo(ws, logo_bytes.read())
        if logo_temp_path:
            temp_files.append(logo_temp_path)

        protect_workbook(wb)

        # Save the workbook to a temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            temp_files.append(tmp.name)  # Add to cleanup list

        # Generate a Brand ID for the customer
        brand_id = generate_brand_id(business_name, email)
        fields['BrandID'] = brand_id

        # Update Notion database with fields and Brand ID
        notion_page_id = update_notion_database(fields=fields)

        # Send the email
        try:
            send_email(
                recipient_email=email,
                subject="Your Custom Invoice Template & Brand ID",
                body="Please find attached your custom Excel invoice template. You can fill in the item details and the calculations will be performed automatically.",
                attachment_paths=[tmp.name],
                business_name=business_name,
                brand_id=brand_id
            )
            
            # Mark this event as successfully processed
            if len(PROCESSED_EVENTS) >= MAX_CACHE_SIZE:
                PROCESSED_EVENTS.popitem(last=False)
            PROCESSED_EVENTS[event_id] = {"timestamp": time.time(), "processed": True}
            
            logger.info(f"✅ Successfully processed event: {event_id}")
            # Then after successfully sending the email, update the Email Sent field
            if notion_page_id:
                try:
                    notion.pages.update(
                        page_id=notion_page_id,
                        properties={
                            "Email Sent": {"checkbox": True}
                        }
                    )
                    logger.info(f"✅ Updated Email Sent status in Notion")
                except Exception as e:
                    logger.error(f"⚠️ Error updating Email Sent status: {e}")
        except Exception as e:
            # Mark as received but not successfully processed
            if len(PROCESSED_EVENTS) >= MAX_CACHE_SIZE:
                PROCESSED_EVENTS.popitem(last=False)
            PROCESSED_EVENTS[event_id] = {"timestamp": time.time(), "processed": False}
            
            logger.error(f"⚠️ Failed to send email: {e}")
            return jsonify({"error": "Failed to send email"}), 500

        return '', 204
        
    except Exception as e:
        # Mark the event as received but not processed if we have an event_id
        if event_id:
            if len(PROCESSED_EVENTS) >= MAX_CACHE_SIZE:
                PROCESSED_EVENTS.popitem(last=False)
            PROCESSED_EVENTS[event_id] = {"timestamp": time.time(), "processed": False}
        
        logger.error(f"⚠️ Unhandled error: {e}")
        return jsonify({"error": str(e)}), 500
        
    finally:
        # Clean up all temporary files
        for file_path in temp_files:
            try:
                if file_path and os.path.exists(file_path):
                    os.unlink(file_path)
                    logger.info(f"✅ Removed temporary file: {file_path}")
            except Exception as e:
                logger.error(f"⚠️ Error removing temporary file {file_path}: {e}")

# === ROUTES ===
@app.route("/health", methods=["GET"])
def health_check():
    """Simple health check endpoint"""
    return jsonify({
        "status": "ok",
        "processed_events": len(PROCESSED_EVENTS),
        "processed_details": {k: v.get("processed", False) for k, v in PROCESSED_EVENTS.items()},
        "timestamp": datetime.now().isoformat()
    })

@app.route("/run-processor", methods=["POST","GET"])
def manual_run():
    """Endpoint to manually trigger the processing job"""
    try:
        process_pending_records()
        return jsonify({
            "status": "success",
            "timestamp": datetime.now().isoformat()
        })
    except Exception as e:
        logger.error(f"Error in manual run: {e}")
        return jsonify({
            "status": "error",
            "error": str(e),
            "timestamp": datetime.now().isoformat()
        }), 500

# === MAIN ===
if __name__ == "__main__":
    # Start the scheduler
    scheduler = start_scheduler()
    
    # Run once at startup
    process_pending_records()
    
    # Start the Flask app
    try:
        app.run(debug=True)
    except (KeyboardInterrupt, SystemExit):
        scheduler.shutdown()
