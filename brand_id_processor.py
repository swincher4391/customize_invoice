from flask import Flask
from notion_client import Client as NotionClient
import requests
import os
import io
import tempfile
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment
import smtplib
from email.message import EmailMessage
from PIL import Image
import time
import hashlib
import uuid
from datetime import datetime, timezone
import logging
from apscheduler.schedulers.background import BackgroundScheduler

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("brandid_processor.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger("BrandIDProcessor")

app = Flask(__name__)

# === ENVIRONMENT VARIABLES ===
NOTION_TOKEN = os.getenv("NOTION_TOKEN")
DATABASE_ID = os.getenv("NOTION_DATABASE_ID")
TEMPLATE_PATH = "invoice-watermarked.xlsx"
SCHEDULER_INTERVAL = int(os.getenv("SCHEDULER_INTERVAL", "60"))  # Default to 60 minutes

# Initialize Notion client
notion = NotionClient(auth=NOTION_TOKEN)

# === UTILITIES ===
def send_email(recipient_email, subject, body, attachment_paths, business_name=''):
    """Send email with attachments and formatted HTML body"""
    smtp_server = os.getenv('SMTP_SERVER', 'smtp.gmail.com')
    smtp_port = int(os.getenv('SMTP_PORT', 587))
    smtp_user = os.getenv('SMTP_USER')
    smtp_pass = os.getenv('SMTP_PASS')
    sender_name = os.getenv('SENDER_NAME', 'Invoice Generator')

    # Create a more sophisticated email
    msg = EmailMessage()
    
    # Better subject line with personalization
    personalized_subject = f"{business_name} - {subject}" if business_name else subject
    msg['Subject'] = personalized_subject
    
    # Add proper From header with sender name
    msg['From'] = f'"{sender_name}" <{smtp_user}>'
    msg['To'] = recipient_email
    
    # Add more headers to improve deliverability
    # Add a unique Message-ID
    domain = smtp_user.split('@')[-1]
    msg['Message-ID'] = f"<{uuid.uuid4()}@{domain}>"
    
    # Add Date header
    msg['Date'] = datetime.now().strftime("%a, %d %b %Y %H:%M:%S %z")
    
    # Add X-Mailer header 
    msg['X-Mailer'] = 'InvoiceCustomizer Service'
    
    # Add a List-Unsubscribe header (helps with spam prevention)
    msg['List-Unsubscribe'] = f'<mailto:{smtp_user}?subject=Unsubscribe>'
    
    # Create personalized HTML content
    brand_id = fields.get('BrandID', 'Not Available')
    
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1">
        <title>{personalized_subject}</title>
        <style>
            body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; padding: 20px; }}
            .container {{ max-width: 600px; margin: 0 auto; }}
            .header {{ background-color: #f8f9fa; padding: 15px; border-bottom: 1px solid #e9ecef; }}
            .content {{ padding: 20px 0; }}
            .footer {{ font-size: 12px; color: #6c757d; padding-top: 20px; border-top: 1px solid #e9ecef; }}
            .brand-id-box {{ background-color: #f0f8ff; border: 1px solid #b0c4de; padding: 15px; border-radius: 5px; margin: 20px 0; text-align: center; }}
            .brand-id {{ font-size: 24px; font-weight: bold; letter-spacing: 1px; color: #0056b3; }}
            .instructions {{ background-color: #fffaf0; border-left: 4px solid #ffa500; padding: 10px; margin: 15px 0; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h2>Your Custom Invoice Template & Brand ID</h2>
            </div>
            <div class="content">
                <p>Hello{' ' + business_name if business_name else ''},</p>
                
                <p>Thank you for using our Invoice Generator service! Your custom Excel invoice template is now ready.</p>
                
                <div class="brand-id-box">
                    <p>Your unique Brand ID is:</p>
                    <p class="brand-id">{brand_id}</p>
                </div>
                
                <div class="instructions">
                    <p><strong>Important:</strong> Save this Brand ID for future purchases. When you buy additional templates on Etsy, include this Brand ID in your order notes to have them automatically customized with your business information.</p>
                </div>
                
                <p>We've attached the spreadsheet to this email. You can fill in the item details, and the calculations will be performed automatically.</p>
                
                <p>This customized invoice includes:</p>
                <ul>
                    <li>Your business information and logo</li>
                    <li>Customized tax rate ({fields.get('Tax %', '')}%)</li>
                    <li>Protected formulas to prevent accidental changes</li>
                    <li>Currency set to {fields.get('Currency', 'USD')}</li>
                </ul>
                
                <p>If you have any questions or need assistance, please reply to this email.</p>
            </div>
            <div class="footer">
                <p>This is a transactional email sent to you because you submitted the Invoice Customization Form.</p>
                <p>To unsubscribe from future emails, please reply with "Unsubscribe" in the subject line.</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    # Set both plain text and HTML content
    msg.set_content(body)
    msg.add_alternative(html_content, subtype='html')

    # Add attachments
    for attachment_path in attachment_paths:
        with open(attachment_path, 'rb') as f:
            file_data = f.read()
            file_name = os.path.basename(attachment_path)

        if attachment_path.endswith('.xlsx'):
            maintype = 'application'
            subtype = 'vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        else:
            maintype = 'application'
            subtype = 'octet-stream'

        msg.add_attachment(file_data, maintype=maintype, subtype=subtype, filename=file_name)

    # Send the email with a retry mechanism
    max_retries = 3
    retry_delay = 2  # seconds
    
    for attempt in range(max_retries):
        try:
            with smtplib.SMTP(smtp_server, smtp_port) as server:
                server.starttls()
                server.login(smtp_user, smtp_pass)
                server.send_message(msg)
                logger.info(f"✅ Email sent successfully to {recipient_email}")
                return True
        except Exception as e:
            if attempt < max_retries - 1:
                logger.warning(f"⚠️ Email attempt {attempt + 1} failed. Retrying in {retry_delay} seconds: {e}")
                time.sleep(retry_delay)
                retry_delay *= 2  # Exponential backoff
            else:
                logger.error(f"❌ All email attempts failed: {e}")
                return False

def remove_background(image_file, tolerance=50):
    """Remove background from logo image"""
    img = Image.open(image_file).convert("RGBA")
    datas = img.getdata()

    new_data = []
    background_color = datas[0]
    logger.debug(f"Detected background color: {background_color}")
    
    for item in datas:
        if all(abs(item[i] - background_color[i]) <= tolerance for i in range(3)):
            new_data.append((255, 255, 255, 0))
        else:
            new_data.append(item)
    img.putdata(new_data)
    return img

def protect_workbook(workbook, password='etsysc123'):
    """Apply protection to Excel workbook"""
    for sheet in workbook.worksheets:
        # Enable protection with specific options
        sheet.protection.sheet = True
        sheet.protection.password = password
        
        # Disable object editing/deletion
        sheet.protection.objects = True
        sheet.protection.scenarios = True
        
        # Other protection options
        sheet.protection.selectLockedCells = False
        sheet.protection.selectUnlockedCells = False
        sheet.protection.formatCells = False
        sheet.protection.formatColumns = False
        sheet.protection.formatRows = False
        sheet.protection.insertColumns = False
        sheet.protection.insertRows = False
        sheet.protection.insertHyperlinks = False
        sheet.protection.deleteColumns = False
        sheet.protection.deleteRows = False
        sheet.protection.sort = False
        sheet.protection.autoFilter = False
        sheet.protection.pivotTables = False
        sheet.protection.drawings = True  # Specifically protects drawings/images

def insert_logo(ws, image_bytes):
    """Insert logo into worksheet"""
    # Create a temporary file that won't be deleted until program exit
    temp_logo = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
    temp_logo.write(image_bytes)
    temp_logo.close()  # Close the file but don't delete it

    try:
        img = OpenpyxlImage(temp_logo.name)
        col_width = ws.column_dimensions['A'].width or 8
        row_height = ws.row_dimensions[1].height or 15

        img.width = int(col_width * 7.5)
        img.height = int(row_height * 1.33)
        img.anchor = 'A1'

        ws.add_image(img)
        
        # Return the filename so we can delete it after the workbook is saved
        return temp_logo.name
    except Exception as e:
        logger.error(f"⚠️ Error adding logo: {e}")
        # If there's an error, try to clean up the file
        try:
            os.unlink(temp_logo.name)
        except:
            pass
        return None

def insert_watermark_background(ws):
    """Insert watermark into worksheet"""
    if os.path.exists("watermark.png"):
        with open("watermark.png", 'rb') as img_file:
            ws._background = img_file.read()

def generate_brand_id(business_name, email):
    """Generate a unique Brand ID following these specific rules:
    
    Format: BRAND-{4-char name code}-{4-digit email code}
    
    Name code algorithm:
    1. For each word, take the first letter to build an initial string
    2. For each word, find first consonant and add after its initial
    3. Limit to 4 characters
    4. Don't repeat letters (use Y instead of X if a second X is needed)
    5. If no consonants, just take the left 4 chars of company name
    
    Examples:
    - "Acme Design Studio" → ACDS (A + C + D + S)
    - "Quick Shop" → QCSH (Q + C + S + H)
    - "AB" → AXBY (A + X + B + Y)
    - "Professional Services" → PRSV (P + R + S + V)
    - "AEIOU" → AEIO (just first 4 chars because no consonants)
    - "XYZ Corp" → XYCR (X + Y + C + R)
    """
    if not business_name or not isinstance(business_name, str) or business_name.strip() == "":
        business_name = "Unknown"
    
    if not email or not isinstance(email, str):
        email = "example@example.com"
    
    # Define vowels for consonant detection
    vowels = ['a', 'e', 'i', 'o', 'u']
    
    # Split into words and filter out empty strings
    words = [word for word in business_name.split() if word]
    if not words:
        words = ["Unknown"]
    
    # Check if there are any consonants in the entire name
    has_consonants = False
    for char in business_name:
        if char.isalpha() and char.lower() not in vowels:
            has_consonants = True
            break
    
    # If no consonants, just take the first 4 characters of the business name
    if not has_consonants:
        name_part = business_name[:4].upper()
        # Pad with X, Y, Z if needed
        if len(name_part) < 4:
            padding_chars = ['X', 'Y', 'Z']
            i = 0
            while len(name_part) < 4:
                name_part += padding_chars[i % 3]
                i += 1
    else:
        # Get initials of each word (first letter)
        initials = ""
        for word in words:
            if word:
                initials += word[0].upper()
        
        # Build result by alternating between initial and first consonant
        name_part = ""
        used_letters = set()  # Track used letters to avoid repetition
        
        for i, word in enumerate(words):
            if len(name_part) >= 4:
                break
                
            # Add the initial if we haven't yet
            if i < len(initials) and initials[i] not in used_letters:
                name_part += initials[i]
                used_letters.add(initials[i])
            
            # Find first consonant in current word
            first_consonant = None
            for char in word:
                if char.isalpha() and char.lower() not in vowels:
                    first_consonant = char.upper()
                    break
            
            # Add first consonant if not already used
            if first_consonant and first_consonant not in used_letters:
                name_part += first_consonant
                used_letters.add(first_consonant)
            elif len(name_part) < 4:
                # No consonant found or already used, add X, Y, or Z
                for padding_char in ['X', 'Y', 'Z']:
                    if padding_char not in used_letters:
                        name_part += padding_char
                        used_letters.add(padding_char)
                        break
            
            # Break if we've reached 4 characters
            if len(name_part) >= 4:
                break
        
        # Trim to 4 characters if longer
        name_part = name_part[:4]
        
        # Pad with unused characters if shorter than 4
        padding_chars = ['X', 'Y', 'Z']
        i = 0
        while len(name_part) < 4:
            padding_char = padding_chars[i % 3]
            if padding_char not in used_letters:
                name_part += padding_char
                used_letters.add(padding_char)
            i += 1
    
    # Calculate ASCII value of email address
    email_ascii_sum = sum(ord(c) for c in email)
    email_part = str(email_ascii_sum)[-4:].zfill(4)  # Get rightmost 4 digits
    
    # Construct final Brand ID
    brand_id = f"BRAND-{name_part}-{email_part}"
    
    logger.info(f"Generated Brand ID for {business_name}: {brand_id}")
    return brand_id
    email_ascii_sum = sum(ord(c) for c in email)
    email_part = str(email_ascii_sum)[-4:].zfill(4)  # Get rightmost 4 digits, pad with leading zeros if needed
    
    # Construct final Brand ID
    brand_id = f"BRAND-{name_part}-{email_part}"
    
    logger.info(f"Generated Brand ID for {business_name}: {brand_id}")
    return brand_id

def extract_notion_properties(page):
    """Extract relevant properties from a Notion page"""
    properties = page.get("properties", {})
    fields = {}
    
    # Extract Company Name
    title_property = next((prop for prop_name, prop in properties.items() 
                         if prop.get("type") == "title"), None)
    if title_property:
        title_objects = properties[title_property["id"]].get("title", [])
        if title_objects:
            fields["Company Name"] = title_objects[0].get("plain_text", "")
    
    # Extract Email
    email_property = next((prop_name for prop_name, prop in properties.items() 
                         if prop.get("type") == "email"), None)
    if email_property:
        fields["Email"] = properties[email_property].get("email", "")
    
    # Extract Logo URL
    for prop_name, prop in properties.items():
        if prop.get("type") == "rich_text" and "logo" in prop_name.lower():
            rich_texts = prop.get("rich_text", [])
            if rich_texts:
                fields["Logo URL"] = rich_texts[0].get("plain_text", "")
    
    # Extract Address
    for prop_name, prop in properties.items():
        if prop.get("type") == "rich_text" and "address" in prop_name.lower():
            rich_texts = prop.get("rich_text", [])
            if rich_texts:
                fields["Address"] = rich_texts[0].get("plain_text", "")
    
    # Extract City, State ZIP
    for prop_name, prop in properties.items():
        if prop.get("type") == "rich_text" and ("city" in prop_name.lower() or "zip" in prop_name.lower()):
            rich_texts = prop.get("rich_text", [])
            if rich_texts:
                fields["City, State ZIP"] = rich_texts[0].get("plain_text", "")
    
    # Extract Phone
    for prop_name, prop in properties.items():
        if prop.get("type") == "phone_number" or (prop.get("type") == "rich_text" and "phone" in prop_name.lower()):
            if prop.get("type") == "phone_number":
                fields["Phone"] = prop.get("phone_number", "")
            else:
                rich_texts = prop.get("rich_text", [])
                if rich_texts:
                    fields["Phone"] = rich_texts[0].get("plain_text", "")
    
    # Extract Tax %
    for prop_name, prop in properties.items():
        if prop.get("type") == "number" and "tax" in prop_name.lower():
            fields["Tax %"] = str(prop.get("number", 7))
    
    # Extract Currency
    for prop_name, prop in properties.items():
        if prop.get("type") == "select" and "currency" in prop_name.lower():
            select = prop.get("select", {})
            fields["Currency"] = select.get("name", "USD")
        elif prop.get("type") == "rich_text" and "currency" in prop_name.lower():
            rich_texts = prop.get("rich_text", [])
            if rich_texts:
                fields["Currency"] = rich_texts[0].get("plain_text", "USD")
    
    return fields

def process_template(fields):
    """Generate a customized template based on customer fields"""
    temp_files = []  # List to keep track of temporary files to clean up
    
    try:
        business_name = fields.get("Company Name", "Your Business")
        address1 = fields.get("Address", "")
        address2 = fields.get("City, State ZIP", "")
        phone = fields.get("Phone", "")
        email = fields.get("Email", "")
        tax_percentage = fields.get("Tax %", "7")
        currency = fields.get("Currency", "USD")
        logo_url = fields.get("Logo URL", "")
        
        if not logo_url:
            logger.error("Logo URL missing!")
            return None, temp_files
        
        # Download and process logo
        logo_response = requests.get(logo_url)
        if logo_response.status_code != 200:
            logger.error(f"Failed to download logo from {logo_url}")
            return None, temp_files
        
        processed_logo = remove_background(io.BytesIO(logo_response.content))
        logo_bytes = io.BytesIO()
        processed_logo.save(logo_bytes, format="PNG")
        logo_bytes.seek(0)
        
        # Load template and customize
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        ws = wb.active
        
        insert_watermark_background(ws)
        
        # Insert business information
        ws['A2'] = business_name
        ws['A3'] = address1
        ws['A4'] = address2
        ws['A5'] = phone
        ws['A6'] = email
        ws['D30'] = f"Tax ({tax_percentage}%)"
        ws['E30'] = f'=IF(NOT(IsGoogleSheets),E29*{float(tax_percentage)}/100,"GOOGLE SHEETS DETECTED")'
        ws['C32'] = f"All amounts shown in {currency}"
        ws.merge_cells('C32:E32')
        ws['C32'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Insert logo
        logo_temp_path = insert_logo(ws, logo_bytes.read())
        if logo_temp_path:
            temp_files.append(logo_temp_path)
        
        # Apply protection
        protect_workbook(wb)
        
        # Save to temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            temp_files.append(tmp.name)
            return tmp.name, temp_files
    
    except Exception as e:
        logger.error(f"Error processing template: {e}")
        return None, temp_files

def update_notion_with_brand_id(page_id, brand_id, email_sent=False):
    """Update Notion record with Brand ID and email status"""
    try:
        properties = {
            "BrandID": {"rich_text": [{"text": {"content": brand_id}}]}
        }
        
        if email_sent:
            properties["Excel Sent"] = {"checkbox": True}
        
        notion.pages.update(
            page_id=page_id,
            properties=properties
        )
        logger.info(f"✅ Updated Notion page {page_id} with Brand ID {brand_id}")
        return True
    except Exception as e:
        logger.error(f"❌ Failed to update Notion page {page_id}: {e}")
        return False

def process_pending_records():
    """Process all Notion records that don't have a Brand ID yet"""
    logger.info("Starting to process pending records...")
    
    try:
        # Query for records without BrandID
        results = notion.databases.query(
            database_id=DATABASE_ID,
            filter={
                "property": "BrandID",
                "rich_text": {
                    "is_empty": True
                }
            }
        ).get("results", [])
        
        logger.info(f"Found {len(results)} records without a Brand ID")
        
        for page in results:
            page_id = page["id"]
            logger.info(f"Processing page {page_id}")
            
            # Extract fields from the Notion page
            fields = extract_notion_properties(page)
            business_name = fields.get("Company Name", "Unknown Business")
            email = fields.get("Email")
            
            if not email:
                logger.warning(f"No email found for page {page_id}, skipping")
                continue
            
            # Generate Brand ID
            brand_id = generate_brand_id(business_name, email)
            fields["BrandID"] = brand_id
            logger.info(f"Generated Brand ID {brand_id} for {business_name}")
            
            # Process template
            template_path, temp_files = process_template(fields)
            if not template_path:
                logger.error(f"Failed to generate template for {business_name}")
                continue
            
            # Send email
            email_success = send_email(
                recipient_email=email,
                subject="Your Custom Invoice Template & Brand ID",
                body=f"Please find attached your custom Excel invoice template. Your Brand ID is: {brand_id}. Please save this ID for future template purchases.",
                attachment_paths=[template_path],
                business_name=business_name
            )
            
            # Update Notion record
            update_notion_with_brand_id(page_id, brand_id, email_success)
            
            # Clean up temporary files
            for file_path in temp_files:
                try:
                    if file_path and os.path.exists(file_path):
                        os.unlink(file_path)
                except Exception as e:
                    logger.warning(f"Failed to remove temporary file {file_path}: {e}")
                    
            # Add a small delay between processing records to avoid rate limits
            time.sleep(1)
        
        logger.info("Finished processing pending records")
    
    except Exception as e:
        logger.error(f"Error in process_pending_records: {e}")

# === SCHEDULER ===
def start_scheduler():
    """Start the background scheduler"""
    scheduler = BackgroundScheduler()
    scheduler.add_job(
        process_pending_records, 
        'interval', 
        minutes=SCHEDULER_INTERVAL,
        id='process_pending_records_job'
    )
    scheduler.start()
    logger.info(f"Scheduler started, will run every {SCHEDULER_INTERVAL} minutes")
    return scheduler

# === HEALTH CHECK ===
@app.route("/health", methods=["GET"])
def health_check():
    """Simple health check endpoint"""
    return {
        "status": "ok",
        "timestamp": datetime.now().isoformat()
    }

# === MANUAL TRIGGER ===
@app.route("/run-processor", methods=["POST"])
def manual_run():
    """Endpoint to manually trigger the processing job"""
    try:
        process_pending_records()
        return {
            "status": "success",
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        logger.error(f"Error in manual run: {e}")
        return {
            "status": "error",
            "error": str(e),
            "timestamp": datetime.now().isoformat()
        }, 500

if __name__ == "__main__":
    # Start the scheduler
    scheduler = start_scheduler()
    
    # Run once at startup
    process_pending_records()
    
    # Start the Flask app
    try:
        app.run(host="0.0.0.0", port=int(os.getenv("PORT", 5000)))
    except (KeyboardInterrupt, SystemExit):
        scheduler.shutdown()
