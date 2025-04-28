from flask import Flask, request, jsonify
from apscheduler.schedulers.background import BackgroundScheduler
from retrying import retry
from requests_oauthlib import OAuth1
from notion_client import Client as NotionClient
import requests
import os
import io
import tempfile
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
import smtplib
from email.message import EmailMessage
from PIL import Image
import time
import re
import json
import logging
from datetime import datetime, timedelta, timezone
from dateutil import parser
import uuid
from collections import OrderedDict

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger('etsy_listener')

app = Flask(__name__)

# === ENVIRONMENT VARIABLES ===
NOTION_TOKEN = os.getenv("NOTION_TOKEN")
BRAND_DATABASE_ID = os.getenv("NOTION_DATABASE_ID")
ORDERS_DATABASE_ID = os.getenv("NOTION_ORDERS_DATABASE_ID", "")  # Add this to your env vars
ETSY_API_KEY = os.getenv("ETSY_API_KEY")
ETSY_API_SECRET = os.getenv("ETSY_API_SECRET")
ETSY_ACCESS_TOKEN = os.getenv("ETSY_ACCESS_TOKEN")
ETSY_ACCESS_TOKEN_SECRET = os.getenv("ETSY_ACCESS_TOKEN_SECRET")
ETSY_SHOP_ID = os.getenv("ETSY_SHOP_ID")
TEMPLATE_PATH = "invoice-template.xlsx"  # Non-watermarked version
WATERMARKED_TEMPLATE_PATH = "invoice-watermarked.xlsx"

# Initialize Notion client
notion = NotionClient(auth=NOTION_TOKEN) if NOTION_TOKEN else None

# Simple in-memory cache of processed events and orders
PROCESSED_EVENTS = OrderedDict()
PROCESSED_ORDERS = OrderedDict()
MAX_CACHE_SIZE = 100

# Global vars to track API health
last_successful_etsy_call = None
last_successful_notion_call = None

# === RETRY LOGIC ===
def retry_if_api_error(exception):
    """Determine if we should retry based on the exception"""
    if isinstance(exception, requests.exceptions.RequestException):
        logger.warning(f"Network error: {str(exception)}")
        return True
    
    if hasattr(exception, 'response') and exception.response is not None:
        status = exception.response.status_code
        if status == 429:  # Rate limit
            logger.warning(f"Rate limited by API: {status}")
            return True
        if 500 <= status < 600:  # Server error
            logger.warning(f"Server error from API: {status}")
            return True
    
    return False

# === ETSY API FUNCTIONS ===
@retry(
    stop_max_attempt_number=3,
    wait_exponential_multiplier=1000,
    wait_exponential_max=10000,
    retry_on_exception=retry_if_api_error
)
def get_etsy_receipts(min_created):
    """Fetch receipts from Etsy API with retry logic"""
    if not all([ETSY_API_KEY, ETSY_API_SECRET, ETSY_ACCESS_TOKEN, ETSY_ACCESS_TOKEN_SECRET, ETSY_SHOP_ID]):
        logger.error("Etsy API credentials not configured")
        return []
    
    # Set up OAuth1 authentication
    auth = OAuth1(
        ETSY_API_KEY,
        client_secret=ETSY_API_SECRET,
        resource_owner_key=ETSY_ACCESS_TOKEN,
        resource_owner_secret=ETSY_ACCESS_TOKEN_SECRET
    )
    
    # API endpoint for shop receipts (orders)
    url = f"https://openapi.etsy.com/v2/shops/{ETSY_SHOP_ID}/receipts"
    params = {
        "min_created": min_created,
        "includes": "Transactions"
    }
    
    logger.info(f"Fetching Etsy receipts created after timestamp {min_created}")
    
    response = requests.get(url, auth=auth, params=params)
    response.raise_for_status()  # Will trigger retry if status code is bad
    
    data = response.json()
    receipts = data.get("results", [])
    
    # Update health tracker
    global last_successful_etsy_call
    last_successful_etsy_call = datetime.now().isoformat()
    
    logger.info(f"Successfully fetched {len(receipts)} receipts from Etsy")
    return receipts

# === TEMPLATE GENERATION FUNCTIONS ===
def remove_background(image_file, tolerance=50):
    """Remove background from logo image"""
    img = Image.open(image_file).convert("RGBA")
    datas = img.getdata()

    new_data = []
    background_color = datas[0]
    logger.debug(f"Detected background color: {background_color}")
    
    for item in datas:
        if all(abs(item[i] - background_color[i]) <= tolerance for i in range(3)):
            new_data.append((255, 255, 255, 0))
        else:
            new_data.append(item)
    img.putdata(new_data)
    return img

def insert_logo(ws, image_bytes):
    """Insert logo into Excel worksheet"""
    # Create a temporary file that won't be deleted until program exit
    temp_logo = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
    temp_logo.write(image_bytes)
    temp_logo.close()  # Close the file but don't delete it

    try:
        img = OpenpyxlImage(temp_logo.name)
        col_width = ws.column_dimensions['A'].width or 8
        row_height = ws.row_dimensions[1].height or 15

        img.width = int(col_width * 7.5)
        img.height = int(row_height * 1.33)
        img.anchor = 'A1'

        ws.add_image(img)
        
        # Return the filename so we can delete it after the workbook is saved
        return temp_logo.name
    except Exception as e:
        logger.error(f"Error adding logo: {e}")
        # If there's an error, try to clean up the file
        try:
            os.unlink(temp_logo.name)
        except:
            pass
        return None

def insert_watermark_background(ws):
    """Insert watermark into worksheet background"""
    if os.path.exists("watermark.png"):
        with open("watermark.png", 'rb') as img_file:
            ws._background = img_file.read()

def protect_workbook(workbook, password='etsysc123'):
    """Apply protection to workbook to prevent formula changes"""
    for sheet in workbook.worksheets:
        # Enable protection with specific options
        sheet.protection.sheet = True
        sheet.protection.password = password
        
        # Disable object editing/deletion
        sheet.protection.objects = True
        sheet.protection.scenarios = True
        
        # Other protection options
        sheet.protection.selectLockedCells = False
        sheet.protection.selectUnlockedCells = False
        sheet.protection.formatCells = False
        sheet.protection.formatColumns = False
        sheet.protection.formatRows = False
        sheet.protection.insertColumns = False
        sheet.protection.insertRows = False
        sheet.protection.insertHyperlinks = False
        sheet.protection.deleteColumns = False
        sheet.protection.deleteRows = False
        sheet.protection.sort = False
        sheet.protection.autoFilter = False
        sheet.protection.pivotTables = False
        sheet.protection.drawings = True  # Specifically protects drawings/images

def generate_excel_template(brand_data, template_path, add_watermark=False):
    """
    Generate Excel template with business branding
    
    Args:
        brand_data: Dictionary with brand information
        template_path: Path to the Excel template
        add_watermark: Whether to add a watermark (for preview)
        
    Returns:
        Path to the generated file
    """
    temp_files = []  # Keep track of temp files to clean up

    try:
        # Extract brand info
        business_name = brand_data.get('Company Name', 'Your Business')
        address1 = brand_data.get('Address', '')
        address2 = brand_data.get('City, State ZIP', '')
        phone = brand_data.get('Phone', '')
        email = brand_data.get('Email', '')
        tax_percentage = brand_data.get('Tax %', '7')
        currency = brand_data.get('Currency', 'USD')
        logo_url = brand_data.get('Logo URL', '')

        if not logo_url:
            logger.error("Logo URL missing in brand data")
            return None, []

        # Download and process logo
        logo_response = requests.get(logo_url)
        if logo_response.status_code != 200:
            logger.error(f"Failed to download logo from {logo_url}")
            return None, []

        processed_logo = remove_background(io.BytesIO(logo_response.content))
        logo_bytes = io.BytesIO()
        processed_logo.save(logo_bytes, format="PNG")
        logo_bytes.seek(0)

        # Load template and customize
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        # Add watermark if needed (preview version)
        if add_watermark:
            insert_watermark_background(ws)

        # Insert business information
        ws['A2'] = business_name
        ws['A3'] = address1
        ws['A4'] = address2
        ws['A5'] = phone
        ws['A6'] = email
        ws['D30'] = f"Tax ({tax_percentage}%)"
        ws['E30'] = f'=IF(NOT(IsGoogleSheets),E29*{float(tax_percentage)}/100,"GOOGLE SHEETS DETECTED")'
        ws['C32'] = f"All amounts shown in {currency}"
        ws.merge_cells('C32:E32')
        
        from openpyxl.styles import Alignment
        ws['C32'].alignment = Alignment(horizontal='center', vertical='center')

        # Insert logo
        logo_temp_path = insert_logo(ws, logo_bytes.read())
        if logo_temp_path:
            temp_files.append(logo_temp_path)

        # Apply protection
        protect_workbook(wb)

        # Save to temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            temp_files.append(tmp.name)
            
        return tmp.name, temp_files
    
    except Exception as e:
        logger.error(f"Error generating Excel template: {e}")
        # Clean up any created temporary files
        for file_path in temp_files:
            try:
                if file_path and os.path.exists(file_path):
                    os.unlink(file_path)
            except Exception as clean_error:
                logger.error(f"Error removing temporary file {file_path}: {clean_error}")
        return None, []

# === EMAIL FUNCTIONS ===
def send_email(recipient_email, subject, body, attachment_paths, business_name=''):
    """
    Send email with attachments
    
    Args:
        recipient_email: Destination email address
        subject: Email subject
        body: Plain text body
        attachment_paths: List of file paths to attach
        business_name: Business name for personalization
    """
    smtp_server = os.getenv('SMTP_SERVER', 'smtp.gmail.com')
    smtp_port = int(os.getenv('SMTP_PORT', 587))
    smtp_user = os.getenv('SMTP_USER')
    smtp_pass = os.getenv('SMTP_PASS')
    sender_name = os.getenv('SENDER_NAME', 'Invoice Generator')

    if not all([smtp_user, smtp_pass]):
        logger.error("SMTP credentials not configured")
        return False

    # Create email message
    msg = EmailMessage()
    
    # Better subject line with personalization
    personalized_subject = f"{business_name} - {subject}" if business_name else subject
    msg['Subject'] = personalized_subject
    msg['From'] = f'"{sender_name}" <{smtp_user}>'
    msg['To'] = recipient_email
    
    # Add headers to improve deliverability
    domain = smtp_user.split('@')[-1]
    msg['Message-ID'] = f"<{uuid.uuid4()}@{domain}>"
    msg['Date'] = datetime.now().strftime("%a, %d %b %Y %H:%M:%S %z")
    msg['X-Mailer'] = 'InvoiceCustomizer Service'
    msg['List-Unsubscribe'] = f'<mailto:{smtp_user}?subject=Unsubscribe>'
    
    # Create personalized HTML content for licensed version
    is_preview = "Preview" in subject
    
    if is_preview:
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <meta name="viewport" content="width=device-width, initial-scale=1">
            <title>{personalized_subject}</title>
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; padding: 20px; }}
                .container {{ max-width: 600px; margin: 0 auto; }}
                .header {{ background-color: #f8f9fa; padding: 15px; border-bottom: 1px solid #e9ecef; }}
                .content {{ padding: 20px 0; }}
                .footer {{ font-size: 12px; color: #6c757d; padding-top: 20px; border-top: 1px solid #e9ecef; }}
                .cta {{ background-color: #0066cc; color: white; padding: 12px 24px; text-decoration: none; border-radius: 4px; display: inline-block; margin: 20px 0; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h2>Your Custom Invoice Preview is Ready!</h2>
                </div>
                <div class="content">
                    <p>Hello{' ' + business_name if business_name else ''},</p>
                    
                    <p>Thank you for requesting a preview of our Invoice Template! The attached preview contains a watermark and shows how your branding looks in the template.</p>
                    
                    <p>This preview includes:</p>
                    <ul>
                        <li>Your business information</li>
                        <li>Your logo</li>
                        <li>A "PREVIEW ONLY" watermark</li>
                    </ul>
                    
                    <p>To get the full, unwatermarked version with all features:</p>
                    
                    <a href="https://www.etsy.com/shop/{ETSY_SHOP_ID}" class="cta">Purchase Full Version on Etsy</a>
                    
                    <p>The licensed version includes:</p>
                    <ul>
                        <li>No watermarks</li>
                        <li>Complete functionality</li>
                        <li>Protected formulas</li>
                        <li>Free updates for 1 year</li>
                    </ul>
                    
                    <p><strong>Important:</strong> When purchasing on Etsy, please include your BrandID in the order notes: <strong>BrandID: {brand_data.get('BrandID', 'Unknown')}</strong></p>
                </div>
                <div class="footer">
                    <p>Questions? Need help? Reply to this email!</p>
                    <p>This is a transactional email sent to you because you submitted the Invoice Customization Form.</p>
                </div>
            </div>
        </body>
        </html>
        """
    else:
        # Licensed version email
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <meta name="viewport" content="width=device-width, initial-scale=1">
            <title>{personalized_subject}</title>
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; padding: 20px; }}
                .container {{ max-width: 600px; margin: 0 auto; }}
                .header {{ background-color: #f8f9fa; padding: 15px; border-bottom: 1px solid #e9ecef; }}
                .content {{ padding: 20px 0; }}
                .footer {{ font-size: 12px; color: #6c757d; padding-top: 20px; border-top: 1px solid #e9ecef; }}
                .license-key {{ background-color: #f8f9fa; padding: 15px; border: 1px solid #e9ecef; margin: 20px 0; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h2>Your Licensed Invoice Template is Ready!</h2>
                </div>
                <div class="content">
                    <p>Hello{' ' + business_name if business_name else ''},</p>
                    
                    <p>Thank you for your purchase! Your fully licensed invoice template is now ready with your custom branding. The template is attached to this email.</p>
                    
                    <div class="license-key">
                        <p><strong>Your License Information:</strong></p>
                        <p>License Key: XLSX-{license_key}</p>
                        <p>Purchase Date: {datetime.now().strftime('%B %d, %Y')}</p>
                        <p>Licensed To: {business_name}</p>
                    </div>
                    
                    <p>Your license entitles you to:</p>
                    <ul>
                        <li>Full use of the template for your business</li>
                        <li>Free updates for 1 year</li>
                        <li>Email support</li>
                    </ul>
                    
                    <p>To get the most from your template:</p>
                    <ol>
                        <li>Open the document in Excel or Google Sheets</li>
                        <li>Fill in your customer information on the right side</li>
                        <li>Add line items in the center section</li>
                        <li>The totals will calculate automatically</li>
                    </ol>
                </div>
                <div class="footer">
                    <p>Need help? Reply to this email for support!</p>
                    <p>Thank you for your business!</p>
                </div>
            </div>
        </body>
        </html>
        """
    
    # Set both plain text and HTML content
    msg.set_content(body)
    msg.add_alternative(html_content, subtype='html')

    # Add attachments
    for attachment_path in attachment_paths:
        with open(attachment_path, 'rb') as f:
            file_data = f.read()
            file_name = os.path.basename(attachment_path)

        if attachment_path.endswith('.xlsx'):
            maintype = 'application'
            subtype = 'vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        else:
            maintype = 'application'
            subtype = 'octet-stream'

        msg.add_attachment(file_data, maintype=maintype, subtype=subtype, filename=file_name)

    # Send the email with retry mechanism
    max_retries = 3
    retry_delay = 2  # seconds
    
    for attempt in range(max_retries):
        try:
            with smtplib.SMTP(smtp_server, smtp_port) as server:
                server.starttls()
                server.login(smtp_user, smtp_pass)
                server.send_message(msg)
                logger.info(f"Email sent successfully to {recipient_email}")
                return True
                
        except Exception as e:
            if attempt < max_retries - 1:
                logger.warning(f"Email attempt {attempt + 1} failed. Retrying in {retry_delay} seconds: {e}")
                time.sleep(retry_delay)
                retry_delay *= 2  # Exponential backoff
            else:
                logger.error(f"All email attempts failed: {e}")
                return False
                
    return False

# === NOTION DATABASE FUNCTIONS ===
def get_brand_data_from_notion(brand_id):
    """Retrieve brand data from Notion using BrandID"""
    if not NOTION_TOKEN or not BRAND_DATABASE_ID:
        logger.error("Notion credentials not set, cannot fetch brand data")
        return None
    
    try:
        # Query Notion for the Brand record
        response = notion.databases.query(
            database_id=BRAND_DATABASE_ID,
            filter={
                "property": "BrandID",
                "rich_text": {
                    "equals": brand_id
                }
            }
        )
        
        # Update health tracker
        global last_successful_notion_call
        last_successful_notion_call = datetime.now().isoformat()
        
        results = response.get("results", [])
        if not results:
            logger.warning(f"No brand data found for BrandID: {brand_id}")
            return None
        
        # Extract the first matching record
        page = results[0]
        properties = page.get("properties", {})
        
        # Map Notion properties to our brand data format
        brand_data = {
            "BrandID": brand_id,
            "Notion_Page_ID": page.get("id")
        }
        
        # Extract properties based on their type
        for prop_name, prop_data in properties.items():
            prop_type = prop_data.get("type")
            
            if prop_type == "title" and prop_data.get("title"):
                brand_data["Company Name"] = prop_data["title"][0]["plain_text"] if prop_data["title"] else ""
            
            elif prop_type == "email":
                brand_data["Email"] = prop_data.get("email", "")
            
            elif prop_type == "phone_number":
                brand_data["Phone"] = prop_data.get("phone_number", "")
            
            elif prop_type == "rich_text" and prop_data.get("rich_text"):
                # Map specific rich_text properties
                if prop_name == "Address":
                    brand_data["Address"] = prop_data["rich_text"][0]["plain_text"] if prop_data["rich_text"] else ""
                elif prop_name == "City, State ZIP":
                    brand_data["City, State ZIP"] = prop_data["rich_text"][0]["plain_text"] if prop_data["rich_text"] else ""
                elif prop_name == "Logo URL":
                    brand_data["Logo URL"] = prop_data["rich_text"][0]["plain_text"] if prop_data["rich_text"] else ""
            
            elif prop_type == "number":
                # Map number properties
                if prop_name == "Tax %":
                    brand_data["Tax %"] = str(prop_data.get("number", 7))
            
            elif prop_type == "select" and prop_data.get("select"):
                # Map select properties
                if prop_name == "Currency":
                    brand_data["Currency"] = prop_data["select"]["name"] if prop_data["select"] else "USD"
        
        logger.info(f"Successfully retrieved brand data for {brand_id}: {brand_data['Company Name']}")
        return brand_data
    
    except Exception as e:
        logger.error(f"Error retrieving brand data from Notion: {e}")
        return None

def update_order_status_in_notion(page_id, status, template_sent=False):
    """Update order status in Notion database"""
    if not NOTION_TOKEN:
        logger.error("Notion credentials not set, cannot update order status")
        return False
    
    try:
        # Update the page with new status
        notion.pages.update(
            page_id=page_id,
            properties={
                "Status": {"select": {"name": status}},
                "Template Sent": {"checkbox": template_sent}
            }
        )
        
        # Update health tracker
        global last_successful_notion_call
        last_successful_notion_call = datetime.now().isoformat()
        
        logger.info(f"Updated order status to '{status}' in Notion")
        return True
    
    except Exception as e:
        logger.error(f"Error updating order status in Notion: {e}")
        return False

def create_order_in_notion(order_id, brand_data, buyer_email):
    """Create new order record in Notion"""
    if not NOTION_TOKEN or not ORDERS_DATABASE_ID:
        logger.error("Notion credentials not set or Orders database ID missing")
        return None
    
    try:
        # Prepare properties for the new order
        properties = {
            "Name": {"title": [{"text": {"content": f"Order #{order_id}"}}]},
            "EtsyOrderID": {"rich_text": [{"text": {"content": str(order_id)}}]},
            "BuyerEmail": {"email": buyer_email},
            "Timestamp": {"date": {"start": datetime.now().isoformat()}},
            "Status": {"select": {"name": "Received"}},
            "Template Sent": {"checkbox": False}
        }
        
        # Add relation to brand if we have a page ID
        if brand_data and brand_data.get("Notion_Page_ID"):
            properties["Brand"] = {"relation": [{"id": brand_data["Notion_Page_ID"]}]}
        
        # Create the page
        response = notion.pages.create(
            parent={"database_id": ORDERS_DATABASE_ID},
            properties=properties
        )
        
        # Update health tracker
        global last_successful_notion_call
        last_successful_notion_call = datetime.now().isoformat()
        
        logger.info(f"Created order #{order_id} in Notion")
        return response["id"]
    
    except Exception as e:
        logger.error(f"Error creating order in Notion: {e}")
        return None

# === MAIN ETSY ORDER PROCESSING FUNCTION ===
def process_etsy_order(receipt):
    """Process a single Etsy order/receipt"""
    order_id = receipt.get("receipt_id")
    
    # Check if we've already processed this order
    if str(order_id) in PROCESSED_ORDERS:
        logger.info(f"Skipping already processed order: {order_id}")
        return
    
    try:
        # Extract order information
        buyer_email = receipt.get("buyer_email")
        message_from_buyer = receipt.get("message_from_buyer", "")
        
        # Extract BrandID using regex
        brand_id_match = re.search(r"BrandID:\s*(\S+)", message_from_buyer)
        brand_id = brand_id_match.group(1) if brand_id_match else None
        
        if not brand_id:
            logger.warning(f"No BrandID found in order #{order_id}, trying to match by email")
            
            # Try to match by email instead
            try:
                response = notion.databases.query(
                    database_id=BRAND_DATABASE_ID,
                    filter={
                        "property": "Email",
                        "email": {
                            "equals": buyer_email
                        }
                    }
                )
                
                results = response.get("results", [])
                if results:
                    # Extract BrandID from the matched record
                    properties = results[0].get("properties", {})
                    for prop_name, prop_data in properties.items():
                        if prop_name == "BrandID" and prop_data.get("rich_text"):
                            brand_id = prop_data["rich_text"][0]["plain_text"]
                            logger.info(f"Matched order #{order_id} to BrandID {brand_id} using email")
                            break
            except Exception as e:
                logger.error(f"Error trying to match by email: {e}")
        
        if not brand_id:
            logger.error(f"Unable to find BrandID for order #{order_id}")
            
            # Add to processed orders to avoid reprocessing
            if len(PROCESSED_ORDERS) >= MAX_CACHE_SIZE:
                PROCESSED_ORDERS.popitem(last=False)
            PROCESSED_ORDERS[str(order_id)] = {"timestamp": time.time(), "processed": False}
            
            return
        
        # Fetch brand data from Notion
        brand_data = get_brand_data_from_notion(brand_id)
        
        if not brand_data:
            logger.error(f"Brand data not found for BrandID {brand_id} (order #{order_id})")
            
            # Add to processed orders to avoid reprocessing
            if len(PROCESSED_ORDERS) >= MAX_CACHE_SIZE:
                PROCESSED_ORDERS.popitem(last=False)
            PROCESSED_ORDERS[str(order_id)] = {"timestamp": time.time(), "processed": False}
            
            return
        
        # Create order record in Notion
        order_page_id = create_order_in_notion(order_id, brand_data, buyer_email)
        
        if not order_page_id:
            logger.error(f"Failed to create order record for #{order_id}")
            return
        
        # Generate the license key
        license_key = f"{order_id}-{uuid.uuid4().hex[:8]}"
        
        # Generate the licensed Excel template
        excel_path, temp_files = generate_excel_template(brand_data, TEMPLATE_PATH, add_watermark=False)
        
        if not excel_path:
            logger.error(f"Failed to generate Excel template for order #{order_id}")
            update_order_status_in_notion(order_page_id, "Failed")
            return
        
        # Send the email with the licensed template
        business_name = brand_data.get("Company Name", "")
        email_sent = send_email(
            recipient_email=buyer_email,
            subject="Your Licensed Invoice Template",
            body=f"Thank you for your purchase! Your licensed invoice template is attached. Your license key is: XLSX-{license_key}",
            attachment_paths=[excel_path],
            business_name=business_name
        )
        
        # Update order status in Notion
        if email_sent:
            update_order_status_in_notion(order_page_id, "Completed", template_sent=True)
            
            # Mark as successfully processed
            if len(PROCESSED_ORDERS) >= MAX_CACHE_SIZE:
                PROCESSED_ORDERS.popitem(last=False)
            PROCESSED_ORDERS[str(order_id)] = {
                "timestamp": time.time(), 
                "processed": True,
                "license_key": license_key
            }
            
            logger.info(f"✅ Successfully processed order #{order_id}")
        else:
            update_order_status_in_notion(order_page_id, "Failed")
            logger.error(f"Failed to send email for order #{order_id}")
        
        # Clean up temporary files
        for file_path in temp_files:
            try:
                if file_path and os.path.exists(file_path):
                    os.unlink(file_path)
                    logger.debug(f"Removed temporary file: {file_path}")
            except Exception as e:
                logger.error(f"Error removing temporary file {file_path}: {e}")
    
    except Exception as e:
        logger.error(f"Error processing order #{order_id}: {e}")
        
        # Mark as failed but processed
        if len(PROCESSED_ORDERS) >= MAX_CACHE_SIZE:
            PROCESSED_ORDERS.popitem(last=False)
        PROCESSED_ORDERS[str(order_id)] = {"timestamp": time.time(), "processed": False}

# === CHECK FOR NEW ETSY ORDERS ===
def check_etsy_orders():
    """
    Poll Etsy API for new receipts/orders and process them
    This function runs periodically via the scheduler
    """
    try:
        # Calculate time window (last 15 minutes)
        min_created = int((datetime.now() - timedelta(minutes=15)).timestamp())
        
        # Fetch receipts from Etsy
        receipts = get_etsy_receipts(min_created)
        
        if receipts:
            logger.info(f"Processing {len(receipts)} new orders from Etsy")
            
            # Process each receipt
            for receipt in receipts:
                process_etsy_order(receipt)
        else:
            logger.info("No new orders found")
    
    except Exception as e:
        logger.error(f"Error in check_etsy_orders: {e}")

# === CHECK FOR ETSY RETURNS/REFUNDS ===
@retry(
    stop_max_attempt_number=3,
    wait_exponential_multiplier=1000,
    wait_exponential_max=10000,
    retry_on_exception=retry_if_api_error
)
def check_etsy_refunds():
    """
    Check for refunded orders and deactivate licenses
    This function runs daily via the scheduler
    """
    if not all([ETSY_API_KEY, ETSY_API_SECRET, ETSY_ACCESS_TOKEN, ETSY_ACCESS_TOKEN_SECRET, ETSY_SHOP_ID]):
        logger.error("Etsy API credentials not configured")
        return
    
    try:
        # Set up OAuth1 authentication
        auth = OAuth1(
            ETSY_API_KEY,
            client_secret=ETSY_API_SECRET,
            resource_owner_key=ETSY_ACCESS_TOKEN,
            resource_owner_secret=ETSY_ACCESS_TOKEN_SECRET
        )
        
        # Calculate time window (last 30 days)
        min_created = int((datetime.now() - timedelta(days=30)).timestamp())
        
        # API endpoint for shop receipts (orders)
        url = f"https://openapi.etsy.com/v2/shops/{ETSY_SHOP_ID}/receipts"
        params = {
            "min_created": min_created,
            "was_paid": True,
            "was_shipped": True
        }
        
        logger.info("Checking for refunded orders")
        
        response = requests.get(url, auth=auth, params=params)
        response.raise_for_status()
        
        receipts = response.json().get("results", [])
        
        refunded_orders = []
        for receipt in receipts:
            if receipt.get("was_refunded", False):
                order_id = receipt.get("receipt_id")
                refunded_orders.append(str(order_id))
        
        if not refunded_orders:
            logger.info("No refunded orders found")
            return
        
        logger.info(f"Found {len(refunded_orders)} refunded orders")
        
        # For each refunded order, update status in Notion
        for order_id in refunded_orders:
            try:
                # Query Notion for the order
                if not ORDERS_DATABASE_ID:
                    logger.error("Orders database ID not set")
                    continue
                
                response = notion.databases.query(
                    database_id=ORDERS_DATABASE_ID,
                    filter={
                        "property": "EtsyOrderID",
                        "rich_text": {
                            "equals": order_id
                        }
                    }
                )
                
                results = response.get("results", [])
                if not results:
                    logger.warning(f"Order #{order_id} not found in Notion")
                    continue
                
                # Update the order status to Refunded
                page_id = results[0]["id"]
                update_order_status_in_notion(page_id, "Refunded", template_sent=True)
                
                # Send notification email about license deactivation
                buyer_email = None
                for prop_name, prop_data in results[0].get("properties", {}).items():
                    if prop_name == "BuyerEmail" and prop_data.get("type") == "email":
                        buyer_email = prop_data.get("email")
                
                if buyer_email:
                    send_email(
                        recipient_email=buyer_email,
                        subject="License Deactivation Notice",
                        body="Your license for our Excel template has been deactivated due to a refund. Please delete all copies of the template. If you believe this is in error, please contact us.",
                        attachment_paths=[],
                        business_name=""
                    )
            
            except Exception as e:
                logger.error(f"Error processing refund for order #{order_id}: {e}")
    
    except Exception as e:
        logger.error(f"Error checking for refunds: {e}")

# === FLASK ROUTES ===
@app.route("/preview_webhook", methods=["POST"])
def handle_preview_request():
    """Handle webhook from Tally form for preview generation"""
    temp_files = []  # List to keep track of temporary files to clean up
    event_id = None
    
    try:
        data = request.json
        logger.info("Raw incoming data from Tally webhook")

        # Extract event ID and creation time for idempotency check
        event_id = data.get('eventId')
        event_time = data.get('createdAt')

        # Check if we've already processed this event SUCCESSFULLY
        if event_id in PROCESSED_EVENTS and PROCESSED_EVENTS[event_id].get("processed", False):
            logger.info(f"Skipping already processed event: {event_id}")
            return '', 204
        
        # If we received it before but didn't process it successfully, try again
        if event_id in PROCESSED_EVENTS:
            logger.info(f"Retrying previously failed event: {event_id}")
        
        # Check if the event is too old (stale)
        is_stale = False
        if event_time:
            try:
                event_datetime = parser.parse(event_time)
                current_time = datetime.now(timezone.utc)
                time_diff = (current_time - event_datetime).total_seconds() / 60
                is_stale = time_diff > 5  # 5 minutes
            except Exception as e:
                logger.error(f"Error parsing event time: {e}")
        
        if is_stale:
            logger.warning(f"Skipping stale event: {event_id}")
            
            # Mark as received but not successfully processed
            if len(PROCESSED_EVENTS) >= MAX_CACHE_SIZE:
                PROCESSED_EVENTS.popitem(last=False)
            PROCESSED_EVENTS[event_id] = {"timestamp": time.time(), "processed": False}
            
            return '', 204

        # Extract fields from the webhook payload
        fields_list = data.get('data', {}).get('fields', [])
        fields = {}

        for field in fields_list:
            label = field.get('label')
            value = field.get('value')

            if isinstance(value, list):
                if value and isinstance(value[0], dict) and 'url' in value[0]:
                    value = value[0]['url']
                else:
                    value = None

            fields[label] = value

        business_name = fields.get('Company Name', 'Your Business')
        address1 = fields.get('Address', '')
        address2 = fields.get('City, State ZIP', '')
        phone = fields.get('Phone', '')
        email = fields.get('Email', '')
        tax_percentage = fields.get('Tax %', '7')
        currency = fields.get('Currency', 'USD')
        logo_url = fields.get('Upload your logo', '')

        if not logo_url:
            logger.error("Logo upload missing!")
            return jsonify({"error": "Logo upload missing!"}), 400
        
        # Generate a unique BrandID
        import hashlib
        brand_id = hashlib.md5(f"{business_name}-{email}-{time.time()}".encode()).hexdigest()[:8].upper()
        fields['BrandID'] = brand_id
        fields['Logo URL'] = logo_url
        
        # Update Notion database with the new BrandID and other fields
        notion_page_id = update_notion_database(fields=fields)
        
        # Generate preview Excel template
        excel_path, temp_files = generate_excel_template(fields, WATERMARKED_TEMPLATE_PATH, add_watermark=True)
        
        if not excel_path:
            logger.error("Failed to generate Excel template")
            return jsonify({"error": "Failed to generate template"}), 500

        # Send the email with preview
        email_sent = send_email(
            recipient_email=email,
            subject="Your Invoice Template Preview",
            body=f"Please find attached your preview. Your BrandID is: {brand_id}. Use this when purchasing the full version.",
            attachment_paths=[excel_path],
            business_name=business_name
        )
        
        if not email_sent:
            logger.error("Failed to send email")
            return jsonify({"error": "Failed to send email"}), 500
            
        # Mark this event as successfully processed
        if len(PROCESSED_EVENTS) >= MAX_CACHE_SIZE:
            PROCESSED_EVENTS.popitem(last=False)
        PROCESSED_EVENTS[event_id] = {"timestamp": time.time(), "processed": True}
        
        logger.info(f"Successfully processed preview request for {business_name} (BrandID: {brand_id})")
        
        # Update the Excel Sent field in Notion
        if notion_page_id:
            try:
                notion.pages.update(
                    page_id=notion_page_id,
                    properties={
                        "Excel Sent": {"checkbox": True},
                        "BrandID": {"rich_text": [{"text": {"content": brand_id}}]}
                    }
                )
                logger.info("Updated Excel Sent status in Notion")
            except Exception as e:
                logger.error(f"Error updating Excel Sent status: {e}")

        return '', 204
        
    except Exception as e:
        # Mark the event as received but not processed if we have an event_id
        if event_id:
            if len(PROCESSED_EVENTS) >= MAX_CACHE_SIZE:
                PROCESSED_EVENTS.popitem(last=False)
            PROCESSED_EVENTS[event_id] = {"timestamp": time.time(), "processed": False}
        
        logger.error(f"Unhandled error in preview webhook: {e}")
        return jsonify({"error": str(e)}), 500
        
    finally:
        # Clean up all temporary files
        for file_path in temp_files:
            try:
                if file_path and os.path.exists(file_path):
                    os.unlink(file_path)
                    logger.debug(f"Removed temporary file: {file_path}")
            except Exception as e:
                logger.error(f"Error removing temporary file {file_path}: {e}")

def update_notion_database(fields, event_id=None):
    """Updates the Notion database with customer information"""
    if not NOTION_TOKEN or not BRAND_DATABASE_ID:
        logger.warning("Notion credentials not set, skipping database update")
        return None
    
    try:
        # Extract relevant information
        business_name = fields.get('Company Name', 'Unknown')
        email = fields.get('Email', '')
        brand_id = fields.get('BrandID', '')
        logo_url = fields.get('Logo URL', '')
        timestamp = datetime.now().isoformat()
        
        # Prepare the data for Notion
        properties = {
            "Name": {"title": [{"text": {"content": business_name}}]},
            "Email": {"email": email},
            "Timestamp": {"date": {"start": timestamp}},
            "BrandID": {"rich_text": [{"text": {"content": brand_id}}]},
            "Logo URL": {"rich_text": [{"text": {"content": logo_url}}]},
            "Validated": {"checkbox": True},
            "Excel Sent": {"checkbox": False}  # Will be updated later
        }
        
        # Add any other fields that exist
        if fields.get('Address'):
            properties["Address"] = {"rich_text": [{"text": {"content": fields.get('Address', '')}}]}
        
        if fields.get('City, State ZIP'):
            properties["City, State ZIP"] = {"rich_text": [{"text": {"content": fields.get('City, State ZIP', '')}}]}
        
        if fields.get('Phone'):
            properties["Phone"] = {"phone_number": fields.get('Phone', '')}
        
        if fields.get('Tax %'):
            properties["Tax %"] = {"number": float(fields.get('Tax %', 7))}
        
        if fields.get('Currency'):
            properties["Currency"] = {"select": {"name": fields.get('Currency', 'USD')}}
        
        # Check if this email already exists in the database
        existing_pages = notion.databases.query(
            database_id=BRAND_DATABASE_ID,
            filter={
                "property": "Email",
                "email": {
                    "equals": email
                }
            }
        ).get("results", [])
        
        if existing_pages:
            # Update existing page
            notion.pages.update(
                page_id=existing_pages[0]["id"],
                properties=properties
            )
            page_id = existing_pages[0]["id"]
            logger.info(f"Updated existing Notion entry for {business_name}")
        else:
            # Create new page
            response = notion.pages.create(
                parent={"database_id": BRAND_DATABASE_ID},
                properties=properties
            )
            page_id = response["id"]
            logger.info(f"Created new Notion entry for {business_name}")
        
        return page_id
            
    except Exception as e:
        logger.error(f"Error updating Notion database: {e}")
        return None

@app.route("/health", methods=["GET"])
def health_check():
    """Enhanced health check endpoint with API status information"""
    
    # Check environment variables are set
    env_vars = {
        "NOTION_TOKEN": os.getenv("NOTION_TOKEN") is not None,
        "NOTION_DATABASE_ID": os.getenv("NOTION_DATABASE_ID") is not None,
        "ETSY_API_KEY": os.getenv("ETSY_API_KEY") is not None,
        "ETSY_SHOP_ID": os.getenv("ETSY_SHOP_ID") is not None,
        "SMTP_SERVER": os.getenv("SMTP_SERVER") is not None,
    }
    
    # Check scheduler status if it exists in the global scope
    scheduler_status = "undefined"
    scheduler_jobs = 0
    next_job_time = None
    
    try:
        if 'scheduler' in globals():
            scheduler_status = "running" if scheduler.running else "stopped"
            scheduler_jobs = len(scheduler.get_jobs())
            if scheduler.get_jobs():
                next_job_time = scheduler.get_jobs()[0].next_run_time.isoformat()
    except Exception as e:
        logger.error(f"Error checking scheduler: {e}")
    
    # Check connectivity to external services
    external_services = {
        "etsy_api": {
            "status": "ok" if last_successful_etsy_call else "unknown",
            "last_success": last_successful_etsy_call
        },
        "notion_api": {
            "status": "ok" if last_successful_notion_call else "unknown",
            "last_success": last_successful_notion_call
        }
    }
    
    # Count processed items
    processed_counts = {
        "events": len(PROCESSED_EVENTS),
        "orders": len(PROCESSED_ORDERS)
    }
    
    return jsonify({
        "status": "ok",
        "timestamp": datetime.now().isoformat(),
        "environment": env_vars,
        "scheduler": {
            "status": scheduler_status,
            "jobs": scheduler_jobs,
            "next_run": next_job_time
        },
        "external_services": external_services,
        "processed_counts": processed_counts
    })

# === INITIALIZE SCHEDULER ===
# Create scheduler
scheduler = BackgroundScheduler()
scheduler.add_job(func=check_etsy_orders, trigger="interval", minutes=5, id="check_orders")
scheduler.add_job(func=check_etsy_refunds, trigger="interval", hours=24, id="check_refunds")

# Initialize function to start the scheduler
def initialize_app():
    """Initialize the application"""
    if not scheduler.running:
        scheduler.start()
        logger.info("Started background scheduler")
    logger.info("Application initialized successfully")

# For Flask 2.2+, use this pattern instead of before_first_request
# Create a blueprint for initialization
from flask import Blueprint

init_blueprint = Blueprint('init_app', __name__)

@init_blueprint.before_app_first_request
def initialize_before_first_request():
    """Initialize before the first request (for Flask 2.2+)"""
    initialize_app()

# Register the blueprint with the app
app.register_blueprint(init_blueprint)

# When running under Gunicorn, we need to initialize immediately
# This is for production deployment
if os.environ.get('GUNICORN_WORKER', '0') == '1':
    logger.info("Running under Gunicorn, initializing scheduler immediately")
    initialize_app()

# Shutdown the scheduler when the app exits
import atexit
atexit.register(lambda: scheduler.shutdown())

if __name__ == "__main__":
    app.run(debug=True, use_reloader=False)  # Important: set use_reloader=False when using APScheduler
